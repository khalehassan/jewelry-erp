from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.contrib import admin
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.test import RequestFactory, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from pypdf import PdfReader

from inventory.models import JewelryItem
from payments.models import Payment
from purchases.models import Purchase, PurchaseLine
from sales.models import Sale, SaleLine
from .models import Account, JournalEntry, JournalLine
from .services import create_entry


class AdminControlTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="admin-controller",
            password="test-password",
            email="admin@example.com",
        )
        self.client.force_login(self.user)
        self.request = RequestFactory().get("/admin/")
        self.request.user = self.user
        self.item = JewelryItem.objects.create(
            name="Protected ring",
            category=JewelryItem.Category.RING,
            karat=JewelryItem.Karat.K21,
            weight_grams=Decimal("2.000"),
            location=JewelryItem.Location.SAFE,
            cost_price=Decimal("1000.00"),
            quantity=1,
        )

    def test_inventory_admin_is_view_only(self):
        item_admin = admin.site._registry[JewelryItem]

        self.assertFalse(item_admin.has_add_permission(self.request))
        self.assertFalse(item_admin.has_change_permission(self.request, self.item))
        self.assertFalse(item_admin.has_delete_permission(self.request, self.item))
        self.assertNotIn("delete_selected", item_admin.get_actions(self.request))

        change_url = reverse("admin:inventory_jewelryitem_change", args=[self.item.pk])
        response = self.client.post(change_url, {
            "name": self.item.name,
            "barcode": self.item.barcode,
            "category": self.item.category,
            "karat": self.item.karat,
            "weight_grams": "2.000",
            "stone_details": "",
            "location": JewelryItem.Location.SHOWCASE,
            "cost_price": "1000.00",
            "quantity": "99",
        })

        self.assertEqual(response.status_code, 403)
        self.item.refresh_from_db()
        self.assertEqual(self.item.location, JewelryItem.Location.SAFE)
        self.assertEqual(self.item.quantity, 1)

    def test_audit_sensitive_admins_have_no_delete_path(self):
        protected_models = (JewelryItem, Purchase, Sale, Payment, JournalEntry, Account)
        for model in protected_models:
            with self.subTest(model=model.__name__):
                model_admin = admin.site._registry[model]
                self.assertFalse(model_admin.has_delete_permission(self.request))
                self.assertNotIn("delete_selected", model_admin.get_actions(self.request))

    def test_existing_journal_entries_are_view_only(self):
        entry = JournalEntry.objects.create(description="Protected journal entry")
        entry_admin = admin.site._registry[JournalEntry]

        self.assertTrue(entry_admin.has_add_permission(self.request))
        self.assertFalse(entry_admin.has_change_permission(self.request, entry))

        change_url = reverse("admin:accounting_journalentry_change", args=[entry.pk])
        delete_url = reverse("admin:accounting_journalentry_delete", args=[entry.pk])
        self.assertEqual(self.client.post(change_url, {
            "date": "2026-08-02",
            "description": "Tampered",
        }).status_code, 403)
        self.assertEqual(self.client.post(delete_url, {"post": "yes"}).status_code, 403)

        entry.refresh_from_db()
        self.assertEqual(entry.description, "Protected journal entry")


class BankMovementReportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="bank-report-owner",
            password="test-password",
        )
        self.user.user_permissions.add(Permission.objects.get(
            content_type__app_label="accounting",
            codename="view_account",
        ))
        self.client.force_login(self.user)

        create_entry(date(2026, 6, 30), "Opening bank funding", [
            ("1021", Decimal("1000.00"), Decimal("0.00")),
            ("3010", Decimal("0.00"), Decimal("1000.00")),
        ])
        create_entry(date(2026, 7, 5), "Bank deposit", [
            ("1021", Decimal("500.00"), Decimal("0.00")),
            ("3010", Decimal("0.00"), Decimal("500.00")),
        ])
        create_entry(date(2026, 7, 10), "Bank withdrawal", [
            ("7550", Decimal("200.00"), Decimal("0.00")),
            ("1021", Decimal("0.00"), Decimal("200.00")),
        ])

    def test_report_has_opening_movements_running_balance_and_match(self):
        response = self.client.get(reverse("accounting:bank_movement"), {
            "account": "1021",
            "from": "2026-07-01",
            "to": "2026-07-31",
            "actual_balance": "1,300.00",
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["opening_balance"], "1,000.00")
        self.assertEqual(response.context["money_in_total"], "500.00")
        self.assertEqual(response.context["money_out_total"], "200.00")
        self.assertEqual(response.context["ending_balance"], "1,300.00")
        self.assertEqual(response.context["actual_balance"], "1,300.00")
        self.assertEqual(response.context["difference"], "0.00")
        self.assertTrue(response.context["is_matched"])
        rows = response.context["rows"]
        self.assertEqual([row["description"] for row in rows], [
            "Bank deposit",
            "Bank withdrawal",
        ])
        self.assertEqual(rows[0]["money_in"], "500.00")
        self.assertEqual(rows[0]["running_balance"], "1,500.00")
        self.assertEqual(rows[1]["money_out"], "200.00")
        self.assertEqual(rows[1]["running_balance"], "1,300.00")

        reports = self.client.get(reverse("accounting:reports"))
        self.assertContains(reports, "Bank Account Movement &amp; Reconciliation")

    def test_report_flags_mismatch_and_invalid_actual_balance(self):
        mismatch = self.client.get(reverse("accounting:bank_movement"), {
            "from": "2026-07-01",
            "to": "2026-07-31",
            "actual_balance": "1250.00",
        })
        self.assertFalse(mismatch.context["is_matched"])
        self.assertTrue(mismatch.context["has_actual_balance"])
        self.assertEqual(mismatch.context["difference"], "-50.00")
        self.assertContains(mismatch, "Bank balance does not match")

        invalid = self.client.get(reverse("accounting:bank_movement"), {
            "from": "2026-07-01",
            "to": "2026-07-31",
            "actual_balance": "not-a-number",
        })
        self.assertFalse(invalid.context["has_actual_balance"])
        self.assertEqual(
            invalid.context["actual_error"],
            "Enter a valid actual bank ending balance.",
        )

    def test_bank_sales_and_purchases_appear_automatically(self):
        purchase = Purchase.objects.create(payment_method=Purchase.PaymentMethod.BANK)
        purchase_line = PurchaseLine.objects.create(
            purchase=purchase,
            name="Bank movement ring",
            karat=21,
            weight_grams=Decimal("1.000"),
            raw_gold_price_per_gram=Decimal("1000.00"),
            craftsmanship_per_gram=Decimal("0.00"),
            stamp_charge=Decimal("0.00"),
            quantity=1,
        )
        purchase.post_to_ledger()
        sale = Sale.objects.create(payment_method=Sale.PaymentMethod.BANK)
        SaleLine.objects.create(
            sale=sale,
            item=purchase_line.created_item,
            gold_price_per_gram=Decimal("1500.00"),
            making_charge_per_gram=Decimal("0.00"),
            quantity=1,
        )
        sale.post_to_ledger()

        today = timezone.localdate().isoformat()
        response = self.client.get(reverse("accounting:bank_movement"), {
            "from": today,
            "to": today,
            "actual_balance": "1800.00",
        })

        rows = response.context["rows"]
        self.assertEqual(len(rows), 2)
        self.assertIn(f"Purchase #{purchase.pk} (Bank)", rows[0]["description"])
        self.assertEqual(rows[0]["money_out"], "1,000.00")
        self.assertIn(f"Sale #{sale.pk} (Bank)", rows[1]["description"])
        self.assertEqual(rows[1]["money_in"], "1,500.00")
        self.assertEqual(response.context["ending_balance"], "1,800.00")
        self.assertTrue(response.context["is_matched"])

    def test_account_selector_and_reversed_dates_are_controlled(self):
        create_entry(timezone.localdate(), "Other payment deposit", [
            ("1025", Decimal("75.00"), Decimal("0.00")),
            ("3010", Decimal("0.00"), Decimal("75.00")),
        ])
        today = timezone.localdate().isoformat()
        response = self.client.get(reverse("accounting:bank_movement"), {
            "account": "1025",
            "from": today,
            "to": today,
            "actual_balance": "75.00",
        })
        self.assertEqual(response.context["account"].code, "1025")
        self.assertEqual(response.context["ending_balance"], "75.00")
        self.assertTrue(response.context["is_matched"])

        reversed_dates = self.client.get(reverse("accounting:bank_movement"), {
            "from": "2026-07-31",
            "to": "2026-07-01",
        })
        self.assertEqual(reversed_dates.context["date_from"], "2026-07-01")
        self.assertEqual(reversed_dates.context["date_to"], "2026-07-31")


class ReportExportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="report-exporter",
            password="test-password",
            first_name="Report",
            last_name="Owner",
        )
        self.user.user_permissions.add(Permission.objects.get(
            content_type__app_label="accounting",
            codename="view_account",
        ))
        self.client.force_login(self.user)
        create_entry(date(2026, 7, 1), "Owner capital deposited", [
            ("1011", Decimal("1000.00"), Decimal("0.00")),
            ("3010", Decimal("0.00"), Decimal("1000.00")),
        ])
        create_entry(date(2026, 7, 2), "Cash deposited to bank", [
            ("1021", Decimal("200.00"), Decimal("0.00")),
            ("1011", Decimal("0.00"), Decimal("200.00")),
        ])

    def _report_urls(self):
        return [
            reverse("accounting:reports"),
            reverse("accounting:trial_balance"),
            reverse("accounting:income_statement"),
            reverse("accounting:balance_sheet"),
            reverse("accounting:inventory_report"),
            reverse("accounting:bank_movement"),
            reverse("accounting:gold_movement"),
            reverse("accounting:account_detail", args=["1011"]),
        ]

    def test_every_report_page_offers_excel_and_pdf_for_current_filters(self):
        for url in self._report_urls():
            with self.subTest(url=url):
                response = self.client.get(url, {
                    "from": "2026-07-01",
                    "to": "2026-07-31",
                })
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, "Download Excel")
                self.assertContains(response, "Download PDF")
                self.assertIn("from=2026-07-01", response.context["excel_export_url"])
                self.assertIn("to=2026-07-31", response.context["excel_export_url"])
                self.assertIn("export=xlsx", response.context["excel_export_url"])
                self.assertIn("export=pdf", response.context["pdf_export_url"])

        reports = self.client.get(reverse("accounting:reports"))
        self.assertContains(reports, "All reports - Excel")
        self.assertContains(reports, "All reports - PDF")

    def test_every_report_downloads_valid_excel_and_pdf(self):
        params = {"from": "2026-07-01", "to": "2026-07-31"}
        for url in self._report_urls():
            with self.subTest(url=url, file_format="xlsx"):
                response = self.client.get(url, {**params, "export": "xlsx"})
                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    response["Content-Type"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                self.assertIn("attachment;", response["Content-Disposition"])
                workbook = load_workbook(BytesIO(response.content), data_only=False)
                self.assertGreaterEqual(len(workbook.sheetnames), 1)
                self.assertTrue(workbook.active["A1"].value)

            with self.subTest(url=url, file_format="pdf"):
                response = self.client.get(url, {**params, "export": "pdf"})
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response["Content-Type"], "application/pdf")
                self.assertTrue(response.content.startswith(b"%PDF-"))
                self.assertGreaterEqual(len(PdfReader(BytesIO(response.content)).pages), 1)

    def test_trial_balance_excel_has_typed_amounts_and_selected_period(self):
        response = self.client.get(reverse("accounting:trial_balance"), {
            "from": "2026-07-01",
            "to": "2026-07-31",
            "show": "all",
            "level": "detail",
            "export": "xlsx",
        })

        workbook = load_workbook(BytesIO(response.content), data_only=False)
        sheet = workbook["Trial Balance"]
        self.assertEqual(sheet["A1"].value, "Trial Balance")
        self.assertEqual(sheet["A2"].value, "2026-07-01 to 2026-07-31")
        cash_row = next(row for row in sheet.iter_rows() if row[0].value == "1011")
        self.assertIsInstance(cash_row[5].value, (int, float))
        self.assertEqual(cash_row[5].value, 1000)
        self.assertEqual(cash_row[6].value, 200)
        self.assertIn("#,##0.00", cash_row[5].number_format)
        self.assertIsNotNone(sheet.freeze_panes)

    def test_bank_pdf_contains_reconciliation_and_filter_details(self):
        response = self.client.get(reverse("accounting:bank_movement"), {
            "account": "1021",
            "from": "2026-07-01",
            "to": "2026-07-31",
            "actual_balance": "200.00",
            "export": "pdf",
        })

        reader = PdfReader(BytesIO(response.content))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        self.assertIn("Bank Account Movement and Reconciliation", text)
        self.assertIn("2026-07-01 to 2026-07-31", text)
        self.assertIn("Cash deposited to bank", text)
        self.assertIn("Matched", text)
        first_page = reader.pages[0]
        self.assertGreater(float(first_page.mediabox.width), float(first_page.mediabox.height))

    def test_excel_treats_user_descriptions_as_text_not_formulas(self):
        create_entry(date(2026, 7, 3), '=HYPERLINK("https://invalid.example","Open")', [
            ("1011", Decimal("1.00"), Decimal("0.00")),
            ("3010", Decimal("0.00"), Decimal("1.00")),
        ])
        response = self.client.get(
            reverse("accounting:account_detail", args=["1011"]),
            {"from": "2026-07-01", "to": "2026-07-31", "export": "xlsx"},
        )

        sheet = load_workbook(BytesIO(response.content), data_only=False).active
        description_cell = next(
            cell
            for row in sheet.iter_rows()
            for cell in row
            if cell.value == '\'=HYPERLINK("https://invalid.example","Open")'
        )
        self.assertEqual(description_cell.data_type, "s")

    def test_complete_report_pack_contains_all_main_reports(self):
        params = {"from": "2026-07-01", "to": "2026-07-31"}
        excel = self.client.get(
            reverse("accounting:export_all_reports", args=["xlsx"]),
            params,
        )
        workbook = load_workbook(BytesIO(excel.content), data_only=False)
        self.assertEqual(workbook.sheetnames, [
            "Reconciliation",
            "Trial Balance",
            "Income Statement",
            "Balance Sheet",
            "Inventory",
            "Bank Movement",
            "Gold Movement",
        ])
        self.assertEqual(workbook["Gold Movement"]["A2"].value, "2026-07-01 to 2026-07-31")

        pdf = self.client.get(
            reverse("accounting:export_all_reports", args=["pdf"]),
            params,
        )
        reader = PdfReader(BytesIO(pdf.content))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        for title in (
            "Operational and Ledger Reconciliation",
            "Trial Balance",
            "Income Statement",
            "Balance Sheet",
            "Inventory Report",
            "Bank Account Movement and Reconciliation",
            "Gold Transaction Log",
        ):
            self.assertIn(title, text)
        self.assertGreaterEqual(len(reader.pages), 7)

    def test_exports_require_permission_and_reject_unknown_formats(self):
        self.client.logout()
        user_without_permission = get_user_model().objects.create_user(
            username="no-report-export",
            password="test-password",
        )
        self.client.force_login(user_without_permission)
        denied = self.client.get(reverse("accounting:trial_balance"), {"export": "xlsx"})
        self.assertRedirects(denied, reverse("sales:dashboard"))
        denied_pack = self.client.get(
            reverse("accounting:export_all_reports", args=["pdf"]),
        )
        self.assertRedirects(denied_pack, reverse("sales:dashboard"))

        self.client.force_login(self.user)
        self.assertEqual(
            self.client.get(reverse("accounting:trial_balance"), {"export": "csv"}).status_code,
            404,
        )
        self.assertEqual(
            self.client.get(reverse("accounting:export_all_reports", args=["csv"])).status_code,
            404,
        )


class GoldMovementReportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="report-owner",
            password="test-password",
        )
        permission = Permission.objects.get(
            content_type__app_label="accounting",
            codename="view_account",
        )
        self.user.user_permissions.add(permission)
        self.client.force_login(self.user)

    def _at(self, year, month, day, hour):
        return timezone.make_aware(datetime(year, month, day, hour, 0))

    def test_purchase_and_sale_are_transaction_rows_with_running_balance(self):
        purchase = Purchase.objects.create()
        line_21k = PurchaseLine.objects.create(
            purchase=purchase,
            name="21K chain",
            karat=21,
            weight_grams=Decimal("2.000"),
            raw_gold_price_per_gram=Decimal("1000.00"),
            craftsmanship_per_gram=Decimal("0.00"),
            stamp_charge=Decimal("0.00"),
            quantity=2,
        )
        Purchase.objects.filter(pk=purchase.pk).update(
            created_at=self._at(2026, 7, 10, 9)
        )
        purchase.refresh_from_db()
        purchase.post_to_ledger()

        sale = Sale.objects.create()
        SaleLine.objects.create(
            sale=sale,
            item=line_21k.created_item,
            gold_price_per_gram=Decimal("3000.00"),
            quantity=1,
        )
        Sale.objects.filter(pk=sale.pk).update(
            created_at=self._at(2026, 7, 10, 11)
        )
        sale.refresh_from_db()
        sale.post_to_ledger()

        response = self.client.get(reverse("accounting:gold_movement"), {
            "from": "2026-07-10",
            "to": "2026-07-10",
        })

        self.assertEqual(response.status_code, 200)
        rows = response.context["rows"]
        self.assertEqual(len(rows), 2)

        self.assertEqual(rows[0]["kind"], "Purchase")
        self.assertEqual(rows[0]["received"], "4.000")
        self.assertEqual(rows[0]["out"], "—")
        self.assertEqual(rows[0]["balance"], "4.000")

        self.assertEqual(rows[1]["kind"], "Sale")
        self.assertEqual(rows[1]["received"], "—")
        self.assertEqual(rows[1]["out"], "2.000")
        self.assertEqual(rows[1]["balance"], "2.000")

        summary_21k = response.context["summaries"][1]
        self.assertEqual(summary_21k["received"], "4.000")
        self.assertEqual(summary_21k["out"], "2.000")
        self.assertEqual(summary_21k["closing"], "2.000")

    def test_purchase_reversal_is_an_outgoing_transaction_and_reconciles(self):
        purchase = Purchase.objects.create()
        PurchaseLine.objects.create(
            purchase=purchase,
            name="Reversed 21K ring",
            karat=21,
            weight_grams=Decimal("2.000"),
            raw_gold_price_per_gram=Decimal("1000.00"),
            craftsmanship_per_gram=Decimal("100.00"),
            stamp_charge=Decimal("50.00"),
            quantity=1,
        )
        purchase.post_to_ledger()
        purchase.reverse(user=self.user, reason="Duplicate purchase")

        today = timezone.localdate().isoformat()
        response = self.client.get(reverse("accounting:gold_movement"), {
            "from": today,
            "to": today,
        })

        rows = response.context["rows"]
        self.assertEqual([row["kind"] for row in rows], [
            "Purchase",
            "Purchase reversal",
        ])
        self.assertEqual(rows[0]["received"], "2.000")
        self.assertEqual(rows[0]["balance"], "2.000")
        self.assertEqual(rows[1]["out"], "2.000")
        self.assertEqual(rows[1]["balance"], "0.000")
        self.assertTrue(response.context["is_reconciled"])

    def test_sale_reversal_is_an_incoming_transaction_and_reconciles(self):
        purchase = Purchase.objects.create()
        purchase_line = PurchaseLine.objects.create(
            purchase=purchase,
            name="Sale reversal stock",
            karat=21,
            weight_grams=Decimal("2.000"),
            raw_gold_price_per_gram=Decimal("1000.00"),
            craftsmanship_per_gram=Decimal("0.00"),
            stamp_charge=Decimal("0.00"),
            quantity=1,
        )
        purchase.post_to_ledger()
        sale = Sale.objects.create()
        SaleLine.objects.create(
            sale=sale,
            item=purchase_line.created_item,
            gold_price_per_gram=Decimal("1500.00"),
            making_charge_per_gram=Decimal("0.00"),
            quantity=1,
        )
        sale.post_to_ledger()
        sale.reverse(user=self.user, reason="Duplicate sale")

        today = timezone.localdate().isoformat()
        response = self.client.get(reverse("accounting:gold_movement"), {
            "from": today,
            "to": today,
        })

        rows = response.context["rows"]
        self.assertEqual([row["kind"] for row in rows], [
            "Purchase",
            "Sale",
            "Sale reversal",
        ])
        self.assertEqual(rows[0]["balance"], "2.000")
        self.assertEqual(rows[1]["out"], "2.000")
        self.assertEqual(rows[1]["balance"], "0.000")
        self.assertEqual(rows[2]["received"], "2.000")
        self.assertEqual(rows[2]["balance"], "2.000")
        self.assertTrue(response.context["is_reconciled"])

    def test_purchase_reversal_after_sale_reversal_closes_gold_and_inventory(self):
        purchase = Purchase.objects.create()
        purchase_line = PurchaseLine.objects.create(
            purchase=purchase,
            name="Fully reversed audit chain",
            karat=21,
            weight_grams=Decimal("2.000"),
            raw_gold_price_per_gram=Decimal("1000.00"),
            craftsmanship_per_gram=Decimal("0.00"),
            stamp_charge=Decimal("0.00"),
            quantity=1,
        )
        purchase.post_to_ledger()
        sale = Sale.objects.create()
        SaleLine.objects.create(
            sale=sale,
            item=purchase_line.created_item,
            gold_price_per_gram=Decimal("1500.00"),
            making_charge_per_gram=Decimal("0.00"),
            quantity=1,
        )
        sale.post_to_ledger()
        sale.reverse(user=self.user, reason="Duplicate sale")
        purchase.reverse(user=self.user, reason="Duplicate purchase")

        today = timezone.localdate().isoformat()
        movement = self.client.get(reverse("accounting:gold_movement"), {
            "from": today,
            "to": today,
        })
        inventory = self.client.get(reverse("accounting:inventory_report"))

        rows = movement.context["rows"]
        self.assertEqual([row["kind"] for row in rows], [
            "Purchase",
            "Sale",
            "Sale reversal",
            "Purchase reversal",
        ])
        self.assertEqual(rows[-1]["balance"], "0.000")
        self.assertTrue(movement.context["is_reconciled"])
        self.assertEqual(inventory.context["total_cost"], "0.00")
        self.assertEqual(inventory.context["ledger_total"], "0.00")
        self.assertTrue(inventory.context["is_reconciled"])
        item = JewelryItem.objects.get(pk=purchase_line.created_item_id)
        self.assertTrue(item.is_archived)
        self.assertEqual(item.quantity, 0)

    def test_filtered_ledger_carries_forward_opening_balance(self):
        opening = Purchase.objects.create(is_opening=True)
        PurchaseLine.objects.create(
            purchase=opening,
            name="Opening 18K gold",
            karat=18,
            weight_grams=Decimal("3.000"),
            raw_gold_price_per_gram=Decimal("500.00"),
            craftsmanship_per_gram=Decimal("0.00"),
            stamp_charge=Decimal("0.00"),
            quantity=1,
        )
        Purchase.objects.filter(pk=opening.pk).update(
            created_at=self._at(2026, 6, 30, 12)
        )
        opening.refresh_from_db()
        opening.post_to_ledger()

        purchase = Purchase.objects.create()
        PurchaseLine.objects.create(
            purchase=purchase,
            name="18K ring",
            karat=18,
            weight_grams=Decimal("2.000"),
            raw_gold_price_per_gram=Decimal("500.00"),
            craftsmanship_per_gram=Decimal("0.00"),
            stamp_charge=Decimal("0.00"),
            quantity=2,
        )
        Purchase.objects.filter(pk=purchase.pk).update(
            created_at=self._at(2026, 7, 1, 9)
        )
        purchase.refresh_from_db()
        purchase.post_to_ledger()

        response = self.client.get(reverse("accounting:gold_movement"), {
            "from": "2026-07-01",
            "to": "2026-07-01",
        })

        rows = response.context["rows"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["kind"], "Purchase")
        self.assertEqual(rows[0]["received"], "4.000")
        self.assertEqual(rows[0]["balance"], "7.000")

        summary_18k = response.context["summaries"][0]
        self.assertEqual(summary_18k["opening"], "3.000")
        self.assertEqual(summary_18k["received"], "4.000")
        self.assertEqual(summary_18k["closing"], "7.000")

    def test_report_requires_accounting_view_permission(self):
        user_without_permission = get_user_model().objects.create_user(
            username="sales-clerk",
            password="test-password",
        )
        self.client.force_login(user_without_permission)

        response = self.client.get(reverse("accounting:gold_movement"))

        self.assertRedirects(response, reverse("sales:dashboard"))


class AutomatedPostingControlTests(TestCase):
    def setUp(self):
        self.cash = Account.objects.get(code="1011")
        self.capital = Account.objects.get(code="3010")

    def test_valid_automated_entry_is_rounded_and_balanced_at_egp_precision(self):
        entry = create_entry(timezone.localdate(), "Valid automated entry", [
            (self.cash.code, Decimal("100.005"), Decimal("0")),
            (self.capital.code, Decimal("0"), Decimal("100.005")),
        ])

        self.assertTrue(entry.is_balanced)
        self.assertEqual(entry.source, JournalEntry.Source.AUTOMATED)
        self.assertIsNone(entry.created_by)
        self.assertEqual(entry.total_debits, Decimal("100.01"))
        self.assertEqual(entry.total_credits, Decimal("100.01"))
        self.assertEqual(entry.lines.count(), 2)

    def test_negative_automated_lines_are_rejected_before_any_write(self):
        invalid_entries = [
            [
                (self.cash.code, Decimal("-100.00"), Decimal("0")),
                (self.capital.code, Decimal("0"), Decimal("-100.00")),
            ],
            [
                (self.cash.code, Decimal("100.00"), Decimal("-1.00")),
                (self.capital.code, Decimal("0"), Decimal("101.00")),
            ],
        ]

        for lines in invalid_entries:
            with self.subTest(lines=lines):
                with self.assertRaisesMessage(ValidationError, "cannot be negative"):
                    create_entry(timezone.localdate(), "Invalid negative entry", lines)

        self.assertEqual(JournalEntry.objects.count(), 0)
        self.assertEqual(JournalLine.objects.count(), 0)

    def test_zero_both_sides_and_nonfinite_lines_are_rejected(self):
        invalid_entries = [
            (
                [
                    (self.cash.code, Decimal("0"), Decimal("0")),
                    (self.capital.code, Decimal("0"), Decimal("0")),
                ],
                "exactly one side",
            ),
            (
                [
                    (self.cash.code, Decimal("100"), Decimal("100")),
                    (self.capital.code, Decimal("100"), Decimal("100")),
                ],
                "exactly one side",
            ),
            (
                [
                    (self.cash.code, Decimal("NaN"), Decimal("0")),
                    (self.capital.code, Decimal("0"), Decimal("1")),
                ],
                "finite monetary amount",
            ),
            (
                [
                    (self.cash.code, Decimal("0.004"), Decimal("0")),
                    (self.capital.code, Decimal("0"), Decimal("0.004")),
                ],
                "rounds to zero",
            ),
        ]

        for lines, message in invalid_entries:
            with self.subTest(lines=lines):
                with self.assertRaisesMessage(ValidationError, message):
                    create_entry(timezone.localdate(), "Invalid line shape", lines)

        self.assertEqual(JournalEntry.objects.count(), 0)

    def test_unbalanced_group_unknown_and_too_short_entries_are_rejected(self):
        group = Account.objects.filter(is_group=True).first()
        invalid_entries = [
            (
                [
                    (self.cash.code, Decimal("100"), Decimal("0")),
                    (self.capital.code, Decimal("0"), Decimal("99")),
                ],
                "not balanced",
            ),
            (
                [
                    (group.code, Decimal("100"), Decimal("0")),
                    (self.capital.code, Decimal("0"), Decimal("100")),
                ],
                "heading, not a postable account",
            ),
            (
                [
                    ("DOES-NOT-EXIST", Decimal("100"), Decimal("0")),
                    (self.capital.code, Decimal("0"), Decimal("100")),
                ],
                "unknown account code",
            ),
            (
                [(self.cash.code, Decimal("100"), Decimal("0"))],
                "at least two posting lines",
            ),
        ]

        for lines, message in invalid_entries:
            with self.subTest(lines=lines):
                with self.assertRaisesMessage(ValidationError, message):
                    create_entry(timezone.localdate(), "Invalid automated entry", lines)

        self.assertEqual(JournalEntry.objects.count(), 0)

    def test_model_and_database_reject_invalid_individual_lines(self):
        entry = JournalEntry.objects.create(description="Constraint test")

        with self.assertRaises(ValidationError):
            JournalLine.objects.create(
                entry=entry,
                account=self.cash,
                debit=Decimal("-1.00"),
                credit=Decimal("0.00"),
            )

        group = Account.objects.filter(is_group=True).first()
        with self.assertRaisesMessage(ValidationError, "heading, not a postable account"):
            JournalLine.objects.create(
                entry=entry,
                account=group,
                debit=Decimal("1.00"),
                credit=Decimal("0.00"),
            )

        invalid_lines = [
            JournalLine(
                entry=entry,
                account=self.cash,
                debit=Decimal("-1.00"),
                credit=Decimal("0.00"),
            ),
            JournalLine(
                entry=entry,
                account=self.cash,
                debit=Decimal("1.00"),
                credit=Decimal("1.00"),
            ),
            JournalLine(
                entry=entry,
                account=self.cash,
                debit=Decimal("0.00"),
                credit=Decimal("0.00"),
            ),
        ]
        for line in invalid_lines:
            with self.subTest(debit=line.debit, credit=line.credit):
                with self.assertRaises(IntegrityError):
                    with transaction.atomic():
                        JournalLine.objects.bulk_create([line])

        self.assertFalse(entry.is_balanced)
        self.assertEqual(entry.lines.count(), 0)


class ManualJournalEntryPageTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="ledger-accountant",
            password="test-password",
        )
        self.user.user_permissions.add(Permission.objects.get(
            content_type__app_label="accounting",
            codename="add_journalentry",
        ))
        self.client.force_login(self.user)
        self.url = reverse("accounting:new_journal_entry")
        self.cash = Account.objects.get(code="1011")
        self.capital = Account.objects.get(code="3010")

    def test_authorized_user_can_open_page_and_only_select_detail_accounts(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        selectable = [
            account
            for group in response.context["account_groups"]
            for account in group["accounts"]
        ]
        self.assertTrue(selectable)
        self.assertTrue(all(not account.is_group for account in selectable))
        self.assertContains(response, "Permanent accounting record")
        self.assertContains(response, "New Journal Entry")

    def test_balanced_manual_entry_posts_with_user_audit_attribution(self):
        response = self.client.post(self.url, {
            "date": "2026-08-03",
            "description": "Owner deposited cash capital",
            "account": [self.cash.code, self.capital.code],
            "debit": ["1000.00", ""],
            "credit": ["", "1000.00"],
        }, follow=True)

        self.assertRedirects(response, self.url)
        entry = JournalEntry.objects.get()
        self.assertEqual(entry.date, date(2026, 8, 3))
        self.assertEqual(entry.description, "Owner deposited cash capital")
        self.assertEqual(entry.source, JournalEntry.Source.MANUAL)
        self.assertEqual(entry.created_by, self.user)
        self.assertTrue(entry.is_balanced)
        self.assertEqual(entry.total_debits, Decimal("1000.00"))
        self.assertContains(response, f"Manual journal entry #{entry.pk} was posted successfully")
        self.assertContains(response, "Owner deposited cash capital")

    def test_unbalanced_or_negative_entry_is_rejected_without_partial_write(self):
        invalid_rows = [
            (["100.00", ""], ["", "99.00"], "not balanced"),
            (["-100.00", ""], ["", "100.00"], "cannot be negative"),
        ]

        for debits, credits, expected_error in invalid_rows:
            with self.subTest(expected_error=expected_error):
                response = self.client.post(self.url, {
                    "date": "2026-08-03",
                    "description": "Invalid manual entry",
                    "account": [self.cash.code, self.capital.code],
                    "debit": debits,
                    "credit": credits,
                })
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, expected_error)
                self.assertContains(response, "Invalid manual entry")
                self.assertEqual(JournalEntry.objects.count(), 0)
                self.assertEqual(JournalLine.objects.count(), 0)

    def test_missing_description_and_incomplete_rows_are_rejected(self):
        response = self.client.post(self.url, {
            "date": "invalid-date",
            "description": "",
            "account": [self.cash.code, ""],
            "debit": ["100.00", ""],
            "credit": ["", "100.00"],
        })

        self.assertContains(response, "Enter a valid journal date")
        self.assertContains(response, "Enter a clear description")
        self.assertContains(response, "Journal line 2: choose an account")
        self.assertEqual(JournalEntry.objects.count(), 0)

    def test_group_account_is_rejected_even_when_posted_directly(self):
        group = Account.objects.filter(is_group=True).first()
        response = self.client.post(self.url, {
            "date": "2026-08-03",
            "description": "Attempted group posting",
            "account": [group.code, self.capital.code],
            "debit": ["100.00", ""],
            "credit": ["", "100.00"],
        })

        self.assertContains(response, "heading, not a postable account")
        self.assertEqual(JournalEntry.objects.count(), 0)

    def test_permission_controls_page_and_navigation(self):
        self.client.logout()
        user_without_permission = get_user_model().objects.create_user(
            username="no-ledger-access",
            password="test-password",
        )
        self.client.force_login(user_without_permission)

        response = self.client.get(self.url)
        self.assertRedirects(response, reverse("sales:dashboard"))
        dashboard = self.client.get(reverse("sales:dashboard"))
        self.assertNotContains(dashboard, "New Journal Entry")

    def test_database_requires_creator_for_manual_source(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                JournalEntry.objects.create(
                    date=date(2026, 8, 3),
                    description="Missing audit user",
                    source=JournalEntry.Source.MANUAL,
                )


class ReportConsistencyTests(TestCase):
    def setUp(self):
        user = get_user_model().objects.create_user(
            username="report-controller",
            password="test-password",
        )
        user.user_permissions.add(Permission.objects.get(
            content_type__app_label="accounting",
            codename="view_account",
        ))
        self.client.force_login(user)

    def test_financial_reports_and_account_ledger_share_the_same_dates(self):
        create_entry(date(2025, 12, 31), "Prior revenue", [
            ("1011", Decimal("100.00"), Decimal("0.00")),
            ("4011", Decimal("0.00"), Decimal("100.00")),
        ])
        create_entry(date(2026, 1, 15), "Current revenue", [
            ("1011", Decimal("200.00"), Decimal("0.00")),
            ("4011", Decimal("0.00"), Decimal("200.00")),
        ])

        period = {"from": "2026-01-01", "to": "2026-12-31"}
        income = self.client.get(reverse("accounting:income_statement"), period)
        trial = self.client.get(reverse("accounting:trial_balance"), {**period, "show": "all"})
        ledger = self.client.get(reverse("accounting:account_detail", args=["4011"]), period)

        self.assertEqual(income.context["revenue_total"], "200.00")
        revenue_trial_row = next(
            row for row in trial.context["rows"] if row["account"].code == "4011"
        )
        self.assertEqual(revenue_trial_row["period_credit"], "200.00")
        self.assertEqual(revenue_trial_row["closing_credit"], "300.00")
        self.assertEqual(ledger.context["opening_balance"], "100.00")
        self.assertEqual(ledger.context["period_credit"], "200.00")
        self.assertEqual(ledger.context["account_balance"], "300.00")

        prior_balance_sheet = self.client.get(
            reverse("accounting:balance_sheet"),
            {"to": "2025-12-31"},
        )
        current_balance_sheet = self.client.get(
            reverse("accounting:balance_sheet"),
            {"to": "2026-12-31"},
        )
        self.assertEqual(prior_balance_sheet.context["asset_total"], "100.00")
        self.assertEqual(prior_balance_sheet.context["total_liab_equity_profit"], "100.00")
        self.assertTrue(prior_balance_sheet.context["is_balanced"])
        self.assertEqual(current_balance_sheet.context["asset_total"], "300.00")
        self.assertEqual(current_balance_sheet.context["total_liab_equity_profit"], "300.00")
        self.assertTrue(current_balance_sheet.context["is_balanced"])

    def test_statements_roll_up_the_trial_balance_account_hierarchy(self):
        create_entry(date(2026, 3, 1), "Gold sales", [
            ("1011", Decimal("1000.00"), Decimal("0.00")),
            ("4011", Decimal("0.00"), Decimal("1000.00")),
        ])
        create_entry(date(2026, 3, 2), "Sales discount", [
            ("4194", Decimal("100.00"), Decimal("0.00")),
            ("1011", Decimal("0.00"), Decimal("100.00")),
        ])
        create_entry(date(2026, 3, 3), "Cost of gold sold", [
            ("5011", Decimal("400.00"), Decimal("0.00")),
            ("1221", Decimal("0.00"), Decimal("400.00")),
        ])
        create_entry(date(2026, 3, 4), "Bank charge", [
            ("6390", Decimal("50.00"), Decimal("0.00")),
            ("1011", Decimal("0.00"), Decimal("50.00")),
        ])

        period = {"from": "2026-03-01", "to": "2026-03-31"}
        income = self.client.get(reverse("accounting:income_statement"), period)
        trial = self.client.get(
            reverse("accounting:trial_balance"),
            {**period, "show": "all"},
        )

        self.assertEqual(income.context["revenue_total"], "900.00")
        self.assertEqual(income.context["expense_total"], "450.00")
        self.assertEqual(income.context["net_profit"], "450.00")
        self.assertEqual(income.context["active_account_count"], 4)

        revenue_sections = {
            section["account"].code: section
            for section in income.context["revenue_sections"]
        }
        self.assertEqual(revenue_sections["4000"]["balance"], "900.00")
        revenue_codes = {
            row["account"].code for row in revenue_sections["4000"]["rows"]
        }
        self.assertTrue({"4010", "4011", "4190", "4194"}.issubset(revenue_codes))

        expense_sections = {
            section["account"].code: section
            for section in income.context["expense_sections"]
        }
        self.assertEqual(expense_sections["5000"]["balance"], "400.00")
        self.assertEqual(expense_sections["6300"]["balance"], "50.00")

        trial_rows = {
            row["account"].code: row for row in trial.context["rows"]
        }
        self.assertEqual(trial_rows["4000"]["closing_credit"], "900.00")
        self.assertEqual(trial_rows["5000"]["closing_debit"], "400.00")
        self.assertEqual(trial_rows["6300"]["closing_debit"], "50.00")

        balance_sheet = self.client.get(
            reverse("accounting:balance_sheet"),
            {"to": "2026-03-31"},
        )
        self.assertEqual(balance_sheet.context["asset_total"], "450.00")
        self.assertEqual(balance_sheet.context["net_profit"], "450.00")
        self.assertEqual(
            balance_sheet.context["total_liab_equity_profit"],
            "450.00",
        )
        self.assertEqual(balance_sheet.context["trial_balance_difference"], "0.00")
        self.assertTrue(balance_sheet.context["is_trial_balanced"])
        self.assertTrue(balance_sheet.context["is_balanced"])

        asset_sections = {
            section["account"].code: section
            for section in balance_sheet.context["asset_sections"]
        }
        self.assertEqual(asset_sections["1000"]["balance"], "850.00")
        self.assertEqual(asset_sections["1200"]["balance"], "-400.00")

    def test_statements_can_show_every_opened_trial_balance_account(self):
        income = self.client.get(reverse("accounting:income_statement"), {
            "from": "2026-01-01",
            "to": "2026-12-31",
            "show": "all",
        })
        balance_sheet = self.client.get(reverse("accounting:balance_sheet"), {
            "to": "2026-12-31",
            "show": "all",
        })

        revenue_roots = {
            section["account"].code
            for section in income.context["revenue_sections"]
        }
        expense_roots = {
            section["account"].code
            for section in income.context["expense_sections"]
        }
        asset_roots = {
            section["account"].code
            for section in balance_sheet.context["asset_sections"]
        }
        liability_roots = {
            section["account"].code
            for section in balance_sheet.context["liability_sections"]
        }
        equity_roots = {
            section["account"].code
            for section in balance_sheet.context["equity_sections"]
        }

        self.assertTrue({"4000", "7000"}.issubset(revenue_roots))
        self.assertTrue({"5000", "6000", "6300", "7500"}.issubset(expense_roots))
        self.assertTrue({"1000", "1200", "1500", "1600"}.issubset(asset_roots))
        self.assertTrue({"2000", "2200"}.issubset(liability_roots))
        self.assertIn("3000", equity_roots)
        self.assertGreater(income.context["detail_account_count"], 0)
        self.assertGreater(balance_sheet.context["detail_account_count"], 0)
        self.assertContains(income, "All opened Trial Balance accounts")
        self.assertContains(balance_sheet, "All opened Trial Balance accounts")

    def test_operational_reports_exclude_unposted_activity_and_show_differences(self):
        posted = Purchase.objects.create()
        PurchaseLine.objects.create(
            purchase=posted,
            name="Posted 18K bracelets",
            category="bracelet",
            karat=18,
            weight_grams=Decimal("2.000"),
            raw_gold_price_per_gram=Decimal("500.00"),
            craftsmanship_per_gram=Decimal("0.00"),
            stamp_charge=Decimal("0.00"),
            quantity=2,
        )
        posted.post_to_ledger()

        today = timezone.localdate().isoformat()
        gold = self.client.get(reverse("accounting:gold_movement"), {
            "from": today,
            "to": today,
        })
        inventory = self.client.get(reverse("accounting:inventory_report"))
        reports = self.client.get(reverse("accounting:reports"))

        self.assertEqual(len(gold.context["rows"]), 1)
        self.assertTrue(gold.context["is_reconciled"])
        self.assertEqual(inventory.context["total_cost"], "2,000.00")
        self.assertEqual(inventory.context["ledger_total"], "2,000.00")
        self.assertTrue(inventory.context["is_reconciled"])
        self.assertTrue(reports.context["reconciliation"]["is_reconciled"])

        unposted = Purchase.objects.create()
        PurchaseLine.objects.create(
            purchase=unposted,
            name="Unposted 21K ring",
            category="ring",
            karat=21,
            weight_grams=Decimal("1.000"),
            raw_gold_price_per_gram=Decimal("500.00"),
            craftsmanship_per_gram=Decimal("0.00"),
            stamp_charge=Decimal("0.00"),
            quantity=1,
        )

        gold = self.client.get(reverse("accounting:gold_movement"), {
            "from": today,
            "to": today,
        })
        inventory = self.client.get(reverse("accounting:inventory_report"))
        reports = self.client.get(reverse("accounting:reports"))

        self.assertEqual(len(gold.context["rows"]), 1)
        self.assertEqual(gold.context["unposted_purchases"], 1)
        self.assertFalse(gold.context["is_reconciled"])
        self.assertEqual(inventory.context["total_cost"], "2,500.00")
        self.assertEqual(inventory.context["ledger_total"], "2,000.00")
        self.assertEqual(inventory.context["difference"], "500.00")
        self.assertFalse(inventory.context["is_reconciled"])
        self.assertEqual(reports.context["reconciliation"]["unposted_total"], 1)
        self.assertFalse(reports.context["reconciliation"]["is_reconciled"])
