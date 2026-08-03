"""Professional Excel and PDF exports for the accounting reports."""

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    LongTable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


@dataclass(frozen=True)
class ExportColumn:
    label: str
    kind: str = "text"
    weight: float = 1


@dataclass(frozen=True)
class ExportRow:
    values: tuple
    style: str = "body"


@dataclass
class ExportTable:
    title: str
    columns: tuple[ExportColumn, ...]
    rows: list[ExportRow] = field(default_factory=list)


@dataclass
class ExportDocument:
    title: str
    subtitle: str
    filename: str
    sheet_name: str
    tables: list[ExportTable]
    metadata: list[tuple[str, object, str]] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    landscape: bool = False


GOLD = "A9812F"
GOLD_LIGHT = "F4EBD5"
DARK = "342D24"
MUTED = "6F6558"
LINE = "DDD4C4"
GREEN = "1E7E34"
GREEN_LIGHT = "E7F4EA"
RED = "B13A32"
RED_LIGHT = "FBE9E7"
WHITE = "FFFFFF"


def _pdf_color(value):
    return colors.HexColor(f"#{value.lstrip('#')}")


def _decimal(value):
    if value in (None, "", "—", "-"):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal(int(value))
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _number(value, default=Decimal("0")):
    parsed = _decimal(value)
    return default if parsed is None else parsed


def _date(value):
    if isinstance(value, (date, datetime)):
        return value
    try:
        return date.fromisoformat(str(value))
    except (TypeError, ValueError):
        return value


def _statement_table(title, sections, total_label, total):
    rows = []
    for section in sections:
        rows.append(ExportRow((
            section["account"].code,
            section["account"].name,
            _number(section["balance"]),
        ), "section"))
        for row in section["rows"]:
            rows.append(ExportRow((
                row["account"].code,
                f"{'  ' * row['depth']}{row['account'].name}",
                _number(row["balance"]),
            ), "group" if row["account"].is_group else "body"))
    if not rows:
        rows.append(ExportRow(("", "No accounts match these filters", None), "muted"))
    rows.append(ExportRow(("", total_label, _number(total)), "total"))
    return ExportTable(
        title,
        (
            ExportColumn("Account code", "text", 1.0),
            ExportColumn("Account", "text", 4.0),
            ExportColumn("Amount (EGP)", "money", 1.5),
        ),
        rows,
    )


def build_report_document(report_key, context):
    if report_key == "reports_overview":
        reconciliation = context["reconciliation"]
        rows = [
            ExportRow((
                check["label"],
                _number(check["operational"]),
                _number(check["ledger"]),
                _number(check["difference"]),
                "Reconciled" if check["is_reconciled"] else "Needs attention",
            ), "body" if check["is_reconciled"] else "warning")
            for check in reconciliation["checks"]
        ]
        unposted = reconciliation["unposted"]
        return ExportDocument(
            title="Operational and Ledger Reconciliation",
            subtitle=f"As of {context['as_of']} - posted transactions only",
            filename=f"operational-ledger-reconciliation-{context['as_of']}",
            sheet_name="Reconciliation",
            metadata=[
                ("Overall status", "Reconciled" if reconciliation["is_reconciled"] else "Needs attention", "status"),
                ("Unposted purchases", unposted["purchases"], "integer"),
                ("Unposted sales", unposted["sales"], "integer"),
                ("Unposted payments", unposted["payments"], "integer"),
            ],
            tables=[ExportTable(
                "Control balances",
                (
                    ExportColumn("Control", "text", 3.0),
                    ExportColumn("Operational (EGP)", "money", 1.5),
                    ExportColumn("Ledger (EGP)", "money", 1.5),
                    ExportColumn("Difference (EGP)", "money", 1.5),
                    ExportColumn("Status", "status", 1.5),
                ),
                rows,
            )],
            notes=["A report is reconciled only when every control difference is zero and all source transactions are posted."],
        )

    if report_key == "trial_balance":
        rows = [
            ExportRow((
                row["account"].code,
                f"{'  ' * row['depth']}{row['account'].name}",
                row["account"].get_type_display(),
                _number(row["opening_debit"]),
                _number(row["opening_credit"]),
                _number(row["period_debit"]),
                _number(row["period_credit"]),
                _number(row["closing_debit"]),
                _number(row["closing_credit"]),
            ), "section" if row["account"].is_group else "body")
            for row in context["rows"]
        ]
        if not rows:
            rows.append(ExportRow(("", "No accounts match these filters", "", None, None, None, None, None, None), "muted"))
        rows.append(ExportRow((
            "", f"Total ({context['row_count']} rows shown)", "",
            None, None, None, None,
            _number(context["total_debit"]),
            _number(context["total_credit"]),
        ), "total"))
        return ExportDocument(
            title="Trial Balance",
            subtitle=f"{context['date_from']} to {context['date_to']}",
            filename=f"trial-balance-{context['date_from']}-to-{context['date_to']}",
            sheet_name="Trial Balance",
            metadata=[
                ("Account level", context["level"].title(), "text"),
                ("Balance type", context["show"].title(), "text"),
                ("Control status", "Balanced" if context["is_balanced"] else "Out of balance", "status"),
                ("Difference (EGP)", _number(context["difference"]), "money"),
            ],
            tables=[ExportTable(
                "Account balances",
                (
                    ExportColumn("Code", "text", 0.9),
                    ExportColumn("Account", "text", 3.2),
                    ExportColumn("Type", "text", 1.1),
                    ExportColumn("Opening debit", "money", 1.2),
                    ExportColumn("Opening credit", "money", 1.2),
                    ExportColumn("Period debit", "money", 1.2),
                    ExportColumn("Period credit", "money", 1.2),
                    ExportColumn("Closing debit", "money", 1.2),
                    ExportColumn("Closing credit", "money", 1.2),
                ),
                rows,
            )],
            notes=["Totals count detail accounts only; group headings are included for presentation and are not counted twice."],
            landscape=True,
        )

    if report_key == "income_statement":
        return ExportDocument(
            title="Income Statement",
            subtitle=f"{context['date_from']} to {context['date_to']}",
            filename=f"income-statement-{context['date_from']}-to-{context['date_to']}",
            sheet_name="Income Statement",
            metadata=[
                ("Accounts displayed", context["show"].title(), "text"),
                ("Active detail accounts", context["active_account_count"], "integer"),
                ("Configured detail accounts", context["detail_account_count"], "integer"),
                ("Net profit / (loss) (EGP)", _number(context["net_profit"]), "money"),
            ],
            tables=[
                _statement_table(
                    "Revenue and other income",
                    context["revenue_sections"],
                    "Total revenue and other income",
                    context["revenue_total"],
                ),
                _statement_table(
                    "Cost of sales and expenses",
                    context["expense_sections"],
                    "Total cost of sales and expenses",
                    context["expense_total"],
                ),
                ExportTable(
                    "Result",
                    (ExportColumn("Measure", "text", 4), ExportColumn("Amount (EGP)", "money", 1.5)),
                    [ExportRow(("Net Profit / (Loss)", _number(context["net_profit"])), "result")],
                ),
            ],
            notes=["Amounts are period movements from the same detail accounts used by the Trial Balance."],
        )

    if report_key == "balance_sheet":
        return ExportDocument(
            title="Balance Sheet",
            subtitle=f"As of {context['as_of']}",
            filename=f"balance-sheet-{context['as_of']}",
            sheet_name="Balance Sheet",
            metadata=[
                ("Balance Sheet status", "Balanced" if context["is_balanced"] else "Needs attention", "status"),
                ("Assets (EGP)", _number(context["asset_total"]), "money"),
                ("Liabilities + equity + profit (EGP)", _number(context["total_liab_equity_profit"]), "money"),
                ("Balance Sheet difference (EGP)", _number(context["difference"]), "money"),
                ("Trial Balance closing debit (EGP)", _number(context["trial_balance_debit"]), "money"),
                ("Trial Balance closing credit (EGP)", _number(context["trial_balance_credit"]), "money"),
                ("Trial Balance difference (EGP)", _number(context["trial_balance_difference"]), "money"),
            ],
            tables=[
                _statement_table("Assets", context["asset_sections"], "Total assets", context["asset_total"]),
                _statement_table("Liabilities", context["liability_sections"], "Total liabilities", context["liability_total"]),
                _statement_table("Equity", context["equity_sections"], "Total equity", context["equity_total"]),
                ExportTable(
                    "Liabilities and equity control",
                    (ExportColumn("Measure", "text", 4), ExportColumn("Amount (EGP)", "money", 1.5)),
                    [
                        ExportRow(("Current-year profit / (loss)", _number(context["net_profit"])), "body"),
                        ExportRow(("Total liabilities + equity + profit", _number(context["total_liab_equity_profit"])), "total"),
                    ],
                ),
            ],
            notes=["Every amount uses the same closing detail-account balances as the Trial Balance."],
        )

    if report_key == "inventory_report":
        summary_rows = [
            ExportRow((
                f"{summary['karat']}K",
                _number(summary["physical"]),
                _number(summary["ledger"]),
                _number(summary["difference"]),
                "Reconciled" if summary["is_reconciled"] else "Difference",
            ), "body" if summary["is_reconciled"] else "warning")
            for summary in context["summaries"]
        ]
        summary_rows.append(ExportRow((
            "Total",
            _number(context["total_cost"]),
            _number(context["ledger_total"]),
            _number(context["difference"]),
            "Reconciled" if context["is_reconciled"] else "Difference",
        ), "total" if context["is_reconciled"] else "warning"))
        detail_rows = []
        for row in context["rows"]:
            item = row["item"]
            detail_rows.append(ExportRow((
                item.barcode,
                item.name,
                item.get_category_display(),
                f"{item.karat}K",
                item.weight_grams,
                item.get_location_display(),
                item.quantity,
                item.cost_price,
                _number(row["line_cost"]),
            )))
        if not detail_rows:
            detail_rows.append(ExportRow(("", "No items in inventory", "", "", None, "", None, None, None), "muted"))
        detail_rows.append(ExportRow(("", "Total stock value (cost)", "", "", None, "", None, None, _number(context["total_cost"])), "total"))
        return ExportDocument(
            title="Inventory Report",
            subtitle=f"As of {context['as_of']}",
            filename=f"inventory-report-{context['as_of']}",
            sheet_name="Inventory",
            metadata=[
                ("Reconciliation status", "Reconciled" if context["is_reconciled"] else "Needs attention", "status"),
                ("Pieces", context["item_count"], "integer"),
                ("Stock records", context["sku_count"], "integer"),
                ("Total stock value (EGP)", _number(context["total_cost"]), "money"),
            ],
            tables=[
                ExportTable(
                    "Physical inventory to ledger",
                    (
                        ExportColumn("Karat", "text", 0.8),
                        ExportColumn("Physical cost (EGP)", "money", 1.4),
                        ExportColumn("Ledger inventory (EGP)", "money", 1.4),
                        ExportColumn("Difference (EGP)", "money", 1.3),
                        ExportColumn("Status", "status", 1.1),
                    ),
                    summary_rows,
                ),
                ExportTable(
                    "Stock details",
                    (
                        ExportColumn("Barcode", "text", 1.3),
                        ExportColumn("Name", "text", 2.3),
                        ExportColumn("Category", "text", 1.2),
                        ExportColumn("Karat", "text", 0.7),
                        ExportColumn("Weight (g)", "weight", 1.0),
                        ExportColumn("Location", "text", 1.1),
                        ExportColumn("Quantity", "integer", 0.8),
                        ExportColumn("Unit cost (EGP)", "money", 1.2),
                        ExportColumn("Cost value (EGP)", "money", 1.3),
                    ),
                    detail_rows,
                ),
            ],
            landscape=True,
        )

    if report_key == "bank_movement":
        metadata = [
            ("Account", f"{context['account'].code} - {context['account'].name}", "text"),
            ("Opening balance (EGP)", _number(context["opening_balance"]), "money"),
            ("Money in (EGP)", _number(context["money_in_total"]), "money"),
            ("Money out (EGP)", _number(context["money_out_total"]), "money"),
            ("Ledger ending balance (EGP)", _number(context["ending_balance"]), "money"),
        ]
        if context["has_actual_balance"]:
            metadata.extend([
                ("Actual bank ending balance (EGP)", _number(context["actual_balance"]), "money"),
                ("Difference (EGP)", _number(context["difference"]), "money"),
                ("Reconciliation status", "Matched" if context["is_matched"] else "Difference", "status"),
            ])
        else:
            metadata.append(("Reconciliation status", "Actual balance not entered", "status"))
        rows = [
            ExportRow((
                row["date"],
                f"Entry #{row['entry_id']}",
                row["description"],
                _decimal(row["money_in"]),
                _decimal(row["money_out"]),
                _number(row["running_balance"]),
            ))
            for row in context["rows"]
        ]
        if not rows:
            rows.append(ExportRow((None, "", "No bank movements in this period", None, None, None), "muted"))
        return ExportDocument(
            title="Bank Account Movement and Reconciliation",
            subtitle=f"{context['date_from']} to {context['date_to']}",
            filename=f"bank-movement-{context['account'].code}-{context['date_from']}-to-{context['date_to']}",
            sheet_name="Bank Movement",
            metadata=metadata,
            tables=[ExportTable(
                "Posted journal movements",
                (
                    ExportColumn("Date", "date", 1.0),
                    ExportColumn("Reference", "text", 1.0),
                    ExportColumn("Description", "text", 3.2),
                    ExportColumn("Money in (EGP)", "money", 1.3),
                    ExportColumn("Money out (EGP)", "money", 1.3),
                    ExportColumn("Running balance (EGP)", "money", 1.5),
                ),
                rows,
            )],
            notes=["Ledger ending balance = opening balance + money in - money out. Posted journal entries only."],
            landscape=True,
        )

    if report_key == "gold_movement":
        summary_columns = [
            ExportColumn("Karat", "text", 0.8),
            ExportColumn("Opening (g)", "weight", 1.1),
            ExportColumn("Received (g)", "weight", 1.1),
            ExportColumn("Out (g)", "weight", 1.1),
            ExportColumn("Closing (g)", "weight", 1.1),
        ]
        if context["show_physical_reconciliation"]:
            summary_columns.extend([
                ExportColumn("Physical now (g)", "weight", 1.2),
                ExportColumn("Difference (g)", "weight", 1.1),
            ])
        summary_rows = []
        for summary in context["summaries"]:
            values = [
                f"{summary['karat']}K",
                _number(summary["opening"]),
                _number(summary["received"]),
                _number(summary["out"]),
                _number(summary["closing"]),
            ]
            if context["show_physical_reconciliation"]:
                values.extend([_number(summary["physical"]), _number(summary["difference"])])
            summary_rows.append(ExportRow(tuple(values), "body" if summary["is_reconciled"] else "warning"))
        movement_rows = [
            ExportRow((
                row["ledger_date"],
                row["occurred_at"],
                row["kind"],
                row["reference"],
                row["party"],
                f"{row['karat']}K",
                _decimal(row["received"]),
                _decimal(row["out"]),
                _number(row["balance"]),
            ))
            for row in context["rows"]
        ]
        if not movement_rows:
            movement_rows.append(ExportRow((None, None, "", "No gold transactions in this period", "", "", None, None, None), "muted"))
        return ExportDocument(
            title="Gold Transaction Log",
            subtitle=f"{context['date_from']} to {context['date_to']}",
            filename=f"gold-transaction-log-{context['date_from']}-to-{context['date_to']}",
            sheet_name="Gold Movement",
            metadata=[
                ("Unposted purchases excluded", context["unposted_purchases"], "integer"),
                ("Unposted sales excluded", context["unposted_sales"], "integer"),
                ("Physical reconciliation", "Reconciled" if context["is_reconciled"] else "Not reconciled / not current date", "status"),
            ],
            tables=[
                ExportTable("Period balance summary", tuple(summary_columns), summary_rows),
                ExportTable(
                    "Transaction movements",
                    (
                        ExportColumn("Ledger date", "date", 1.0),
                        ExportColumn("Entered at", "datetime", 1.3),
                        ExportColumn("Transaction", "text", 1.2),
                        ExportColumn("Reference", "text", 1.5),
                        ExportColumn("Party / payment / reason", "text", 2.5),
                        ExportColumn("Karat", "text", 0.7),
                        ExportColumn("Received (g)", "weight", 1.1),
                        ExportColumn("Out (g)", "weight", 1.0),
                        ExportColumn("Running balance (g)", "weight", 1.3),
                    ),
                    movement_rows,
                ),
            ],
            notes=["Balances are maintained separately for 18K, 21K, and 24K gold and use posted transactions only."],
            landscape=True,
        )

    if report_key == "account_detail":
        account = context["account"]
        rows = [
            ExportRow((
                row["line"].entry.date,
                f"Entry #{row['line'].entry_id}",
                row["line"].entry.description or "Entry",
                _number(row["debit"]),
                _number(row["credit"]),
                _number(row["running"]),
            ))
            for row in context["rows"]
        ]
        if not rows:
            rows.append(ExportRow((None, "", "No transactions for this account in the period", None, None, None), "muted"))
        return ExportDocument(
            title=f"{account.code} - {account.name} Ledger",
            subtitle=f"{context['date_from']} to {context['date_to']}",
            filename=f"account-{account.code}-ledger-{context['date_from']}-to-{context['date_to']}",
            sheet_name=f"{account.code} Ledger",
            metadata=[
                ("Account type", account.get_type_display(), "text"),
                ("Opening balance (EGP)", _number(context["opening_balance"]), "money"),
                ("Period debits (EGP)", _number(context["period_debit"]), "money"),
                ("Period credits (EGP)", _number(context["period_credit"]), "money"),
                ("Closing balance (EGP)", _number(context["account_balance"]), "money"),
            ],
            tables=[ExportTable(
                "Journal transactions",
                (
                    ExportColumn("Date", "date", 1.0),
                    ExportColumn("Reference", "text", 1.0),
                    ExportColumn("Description", "text", 3.0),
                    ExportColumn("Debit (EGP)", "money", 1.2),
                    ExportColumn("Credit (EGP)", "money", 1.2),
                    ExportColumn("Balance (EGP)", "money", 1.3),
                ),
                rows,
            )],
            landscape=True,
        )

    raise ValueError(f"Unknown report export key: {report_key}")


def _excel_value(value, kind):
    if value is None:
        return None
    if kind in ("money", "weight"):
        number = _decimal(value)
        return float(number) if number is not None else None
    if kind == "integer":
        number = _decimal(value)
        return int(number) if number is not None else None
    if kind in ("date", "datetime"):
        parsed = _date(value)
        if isinstance(parsed, datetime) and timezone.is_aware(parsed):
            parsed = timezone.localtime(parsed).replace(tzinfo=None)
        return parsed
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _excel_number_format(kind):
    return {
        "money": '#,##0.00;[Red](#,##0.00);-',
        "weight": '#,##0.000;[Red](#,##0.000);-',
        "integer": '#,##0',
        "date": 'yyyy-mm-dd',
        "datetime": 'yyyy-mm-dd hh:mm',
        "text": '@',
        "status": '@',
    }.get(kind, "General")


def _apply_excel_row_style(cells, style):
    if style == "section":
        for cell in cells:
            cell.fill = PatternFill("solid", fgColor=GOLD_LIGHT)
            cell.font = Font(bold=True, color=DARK)
    elif style == "group":
        for cell in cells:
            cell.font = Font(bold=True, color=DARK)
    elif style == "total":
        for cell in cells:
            cell.font = Font(bold=True, color=DARK)
            cell.border = Border(top=Side(style="medium", color=GOLD))
    elif style == "result":
        for cell in cells:
            cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
            cell.font = Font(bold=True, size=12, color=GREEN)
            cell.border = Border(top=Side(style="double", color=GOLD))
    elif style == "warning":
        for cell in cells:
            cell.fill = PatternFill("solid", fgColor=RED_LIGHT)
            cell.font = Font(color=RED)
    elif style == "muted":
        for cell in cells:
            cell.font = Font(italic=True, color=MUTED)


def _unique_sheet_name(workbook, requested):
    base = requested[:31] or "Report"
    candidate = base
    suffix = 2
    while candidate in workbook.sheetnames:
        suffix_text = f" {suffix}"
        candidate = f"{base[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    return candidate


def _build_xlsx(documents, generated_by, generated_at):
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Jewelry ERP"
    workbook.properties.title = documents[0].title if len(documents) == 1 else "Jewelry ERP Accounting Report Pack"
    workbook.properties.created = generated_at.replace(tzinfo=None)

    for document in documents:
        sheet = workbook.create_sheet(_unique_sheet_name(workbook, document.sheet_name))
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape" if document.landscape else "portrait"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.oddFooter.center.text = "Page &P of &N"
        sheet.oddFooter.right.text = "Jewelry ERP"

        max_columns = max(
            [2] + [len(table.columns) for table in document.tables]
        )
        final_column = get_column_letter(max_columns)
        sheet.merge_cells(f"A1:{final_column}1")
        title_cell = sheet["A1"]
        title_cell.value = document.title
        title_cell.fill = PatternFill("solid", fgColor=DARK)
        title_cell.font = Font(bold=True, size=18, color=WHITE)
        title_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 30

        sheet.merge_cells(f"A2:{final_column}2")
        sheet["A2"] = document.subtitle
        sheet["A2"].font = Font(italic=True, color=MUTED)
        sheet["A2"].alignment = Alignment(wrap_text=True)

        sheet.merge_cells(f"A3:{final_column}3")
        sheet["A3"] = f"Generated {generated_at:%Y-%m-%d %H:%M} by {generated_by}"
        sheet["A3"].font = Font(size=9, color=MUTED)

        row_number = 5
        column_widths = [10.0] * max_columns
        for label, value, kind in document.metadata:
            label_cell = sheet.cell(row_number, 1, label)
            value_cell = sheet.cell(row_number, 2, _excel_value(value, kind))
            label_cell.font = Font(bold=True, color=MUTED)
            value_cell.number_format = _excel_number_format(kind)
            value_cell.alignment = Alignment(horizontal="right" if kind in ("money", "weight", "integer") else "left")
            if kind == "status":
                status_text = str(value).lower()
                color = GREEN if any(word in status_text for word in ("balanced", "matched", "reconciled")) and "not " not in status_text else RED
                value_cell.font = Font(bold=True, color=color)
            column_widths[0] = max(column_widths[0], min(38.0, len(str(label)) + 2.0))
            column_widths[1] = max(column_widths[1], min(42.0, len(str(value)) + 2.0))
            row_number += 1

        row_number += 1
        first_table_header = None
        for table in document.tables:
            sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=len(table.columns))
            section_cell = sheet.cell(row_number, 1, table.title)
            section_cell.fill = PatternFill("solid", fgColor=GOLD)
            section_cell.font = Font(bold=True, color=WHITE, size=11)
            section_cell.alignment = Alignment(vertical="center")
            sheet.row_dimensions[row_number].height = 22
            row_number += 1

            header_row = row_number
            if first_table_header is None:
                first_table_header = header_row
            for column_number, column in enumerate(table.columns, start=1):
                cell = sheet.cell(row_number, column_number, column.label)
                cell.fill = PatternFill("solid", fgColor=DARK)
                cell.font = Font(bold=True, color=WHITE)
                cell.alignment = Alignment(
                    horizontal="right" if column.kind in ("money", "weight", "integer") else "left",
                    vertical="center",
                    wrap_text=True,
                )
                column_widths[column_number - 1] = max(
                    column_widths[column_number - 1],
                    min(34.0, max(11.0, column.weight * 7.0)),
                )
            sheet.row_dimensions[row_number].height = 42
            row_number += 1

            for export_row in table.rows:
                excel_cells = []
                for column_number, (column, value) in enumerate(zip(table.columns, export_row.values), start=1):
                    cell = sheet.cell(row_number, column_number, _excel_value(value, column.kind))
                    cell.number_format = _excel_number_format(column.kind)
                    cell.alignment = Alignment(
                        horizontal="right" if column.kind in ("money", "weight", "integer") else "left",
                        vertical="top",
                        wrap_text=column.kind in ("text", "status"),
                    )
                    cell.border = Border(bottom=Side(style="hair", color=LINE))
                    excel_cells.append(cell)
                    display_length = len(str(value or ""))
                    column_widths[column_number - 1] = max(
                        column_widths[column_number - 1],
                        min(42.0, display_length + 2.0),
                    )
                _apply_excel_row_style(excel_cells, export_row.style)
                row_number += 1

            row_number += 2

        if document.notes:
            sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=max_columns)
            sheet.cell(row_number, 1, "Notes")
            sheet.cell(row_number, 1).font = Font(bold=True, color=GOLD)
            row_number += 1
            for note in document.notes:
                sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=max_columns)
                cell = sheet.cell(row_number, 1, note)
                cell.font = Font(size=9, italic=True, color=MUTED)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                sheet.row_dimensions[row_number].height = 28
                row_number += 1

        for column_number, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(column_number)].width = min(width, 42)
        if first_table_header:
            sheet.freeze_panes = f"A{first_table_header + 1}"
        sheet.print_area = f"A1:{final_column}{max(1, row_number)}"

    workbook.active = 0
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _register_pdf_fonts():
    candidates = [
        (
            Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
            Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
        ),
        (
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ),
        (
            Path("/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf"),
            Path("/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf"),
        ),
    ]
    for regular, bold in candidates:
        if regular.exists() and bold.exists():
            try:
                pdfmetrics.registerFont(TTFont("ERPReport", str(regular)))
                pdfmetrics.registerFont(TTFont("ERPReportBold", str(bold)))
                return "ERPReport", "ERPReportBold"
            except Exception:
                continue
    return "Helvetica", "Helvetica-Bold"


def _pdf_text(value):
    if value is None:
        return "-"
    return (
        str(value)
        .replace("—", "-")
        .replace("−", "-")
        .replace("·", "-")
        .replace("✓", "PASS")
        .replace("⚠", "WARNING")
    )


def _display_value(value, kind):
    if value is None:
        return "-"
    if kind == "money":
        number = _decimal(value)
        return f"{number:,.2f}" if number is not None else _pdf_text(value)
    if kind == "weight":
        number = _decimal(value)
        return f"{number:,.3f}" if number is not None else _pdf_text(value)
    if kind == "integer":
        number = _decimal(value)
        return f"{int(number):,}" if number is not None else _pdf_text(value)
    if kind == "date":
        parsed = _date(value)
        return parsed.strftime("%Y-%m-%d") if hasattr(parsed, "strftime") else _pdf_text(parsed)
    if kind == "datetime":
        parsed = _date(value)
        if isinstance(parsed, datetime):
            if timezone.is_aware(parsed):
                parsed = timezone.localtime(parsed)
            return parsed.strftime("%Y-%m-%d %H:%M")
        return _pdf_text(parsed)
    return _pdf_text(value)


def _build_pdf(documents, generated_by, generated_at):
    body_font, bold_font = _register_pdf_fonts()
    pagesize = landscape(A4) if len(documents) > 1 or any(document.landscape for document in documents) else A4
    output = BytesIO()
    pdf = SimpleDocTemplate(
        output,
        pagesize=pagesize,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=15 * mm,
        title=documents[0].title if len(documents) == 1 else "Jewelry ERP Accounting Report Pack",
        author="Jewelry ERP",
    )
    base_styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ERPTitle",
        parent=base_styles["Title"],
        fontName=bold_font,
        fontSize=18,
        leading=22,
        textColor=_pdf_color(DARK),
        alignment=TA_LEFT,
        spaceAfter=4 * mm,
    )
    subtitle_style = ParagraphStyle(
        "ERPSubtitle",
        parent=base_styles["Normal"],
        fontName=body_font,
        fontSize=9,
        leading=12,
        textColor=_pdf_color(MUTED),
        spaceAfter=3 * mm,
    )
    table_title_style = ParagraphStyle(
        "ERPTableTitle",
        parent=base_styles["Heading2"],
        fontName=bold_font,
        fontSize=10,
        leading=12,
        textColor=_pdf_color(GOLD),
        spaceBefore=3 * mm,
        spaceAfter=2 * mm,
        keepWithNext=1,
    )
    text_style = ParagraphStyle(
        "ERPTableText",
        parent=base_styles["Normal"],
        fontName=body_font,
        fontSize=7.2,
        leading=9,
        textColor=_pdf_color(DARK),
        wordWrap="CJK",
    )
    number_style = ParagraphStyle(
        "ERPTableNumber",
        parent=text_style,
        alignment=TA_RIGHT,
    )
    note_style = ParagraphStyle(
        "ERPNote",
        parent=base_styles["Normal"],
        fontName=body_font,
        fontSize=7.5,
        leading=10,
        textColor=_pdf_color(MUTED),
        spaceBefore=2 * mm,
    )

    story = []
    for document_index, document in enumerate(documents):
        if document_index:
            story.append(PageBreak())
        story.append(Paragraph(escape(_pdf_text(document.title)), title_style))
        story.append(Paragraph(escape(_pdf_text(document.subtitle)), subtitle_style))
        story.append(Paragraph(
            escape(f"Generated {generated_at:%Y-%m-%d %H:%M} by {generated_by}"),
            subtitle_style,
        ))

        if document.metadata:
            metadata_data = []
            for label, value, kind in document.metadata:
                metadata_data.append([
                    Paragraph(escape(_pdf_text(label)), text_style),
                    Paragraph(escape(_display_value(value, kind)), number_style if kind in ("money", "weight", "integer") else text_style),
                ])
            metadata_table = Table(metadata_data, colWidths=[pdf.width * .42, pdf.width * .53], hAlign="LEFT")
            metadata_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (0, -1), bold_font),
                ("TEXTCOLOR", (0, 0), (0, -1), _pdf_color(MUTED)),
                ("BACKGROUND", (0, 0), (-1, -1), _pdf_color("F7F3EA")),
                ("BOX", (0, 0), (-1, -1), .5, _pdf_color(LINE)),
                ("INNERGRID", (0, 0), (-1, -1), .25, _pdf_color(LINE)),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.extend([metadata_table, Spacer(1, 3 * mm)])

        for export_table in document.tables:
            story.append(Paragraph(escape(_pdf_text(export_table.title)), table_title_style))
            header = [
                Paragraph(escape(_pdf_text(column.label)), ParagraphStyle(
                    f"Header{column_index}",
                    parent=text_style,
                    fontName=bold_font,
                    textColor=colors.white,
                    alignment=TA_RIGHT if column.kind in ("money", "weight", "integer") else TA_LEFT,
                ))
                for column_index, column in enumerate(export_table.columns)
            ]
            data = [header]
            for row in export_table.rows:
                cells = []
                for column, value in zip(export_table.columns, row.values):
                    style = number_style if column.kind in ("money", "weight", "integer") else text_style
                    cells.append(Paragraph(escape(_display_value(value, column.kind)), style))
                data.append(cells)

            total_weight = sum(column.weight for column in export_table.columns)
            col_widths = [pdf.width * .98 * column.weight / total_weight for column in export_table.columns]
            table = LongTable(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT", splitByRow=1)
            commands = [
                ("BACKGROUND", (0, 0), (-1, 0), _pdf_color(DARK)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LINEBELOW", (0, 0), (-1, -1), .25, _pdf_color(LINE)),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
            for row_index, export_row in enumerate(export_table.rows, start=1):
                if export_row.style == "section":
                    commands.extend([
                        ("BACKGROUND", (0, row_index), (-1, row_index), _pdf_color(GOLD_LIGHT)),
                        ("FONTNAME", (0, row_index), (-1, row_index), bold_font),
                    ])
                elif export_row.style == "group":
                    commands.append(("FONTNAME", (0, row_index), (-1, row_index), bold_font))
                elif export_row.style == "total":
                    commands.extend([
                        ("FONTNAME", (0, row_index), (-1, row_index), bold_font),
                        ("LINEABOVE", (0, row_index), (-1, row_index), 1, _pdf_color(GOLD)),
                    ])
                elif export_row.style == "result":
                    commands.extend([
                        ("BACKGROUND", (0, row_index), (-1, row_index), _pdf_color(GREEN_LIGHT)),
                        ("TEXTCOLOR", (0, row_index), (-1, row_index), _pdf_color(GREEN)),
                        ("FONTNAME", (0, row_index), (-1, row_index), bold_font),
                        ("LINEABOVE", (0, row_index), (-1, row_index), 1.5, _pdf_color(GOLD)),
                    ])
                elif export_row.style == "warning":
                    commands.extend([
                        ("BACKGROUND", (0, row_index), (-1, row_index), _pdf_color(RED_LIGHT)),
                        ("TEXTCOLOR", (0, row_index), (-1, row_index), _pdf_color(RED)),
                    ])
                elif export_row.style == "muted":
                    commands.append(("TEXTCOLOR", (0, row_index), (-1, row_index), _pdf_color(MUTED)))
            table.setStyle(TableStyle(commands))
            story.extend([table, Spacer(1, 3 * mm)])

        for note in document.notes:
            story.append(Paragraph(f"<b>Note:</b> {escape(_pdf_text(note))}", note_style))

    def draw_footer(canvas, document_template):
        canvas.saveState()
        canvas.setFont(body_font, 7)
        canvas.setFillColor(_pdf_color(MUTED))
        canvas.drawString(12 * mm, 7 * mm, "Jewelry ERP - Confidential accounting report")
        canvas.drawRightString(pagesize[0] - 12 * mm, 7 * mm, f"Page {document_template.page}")
        canvas.restoreState()

    pdf.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    return output.getvalue()


def export_documents_response(documents, file_format, generated_by, filename=None):
    if file_format not in ("xlsx", "pdf"):
        raise ValueError("Export format must be xlsx or pdf.")
    generated_at = timezone.localtime()
    base_filename = filename or documents[0].filename
    if file_format == "xlsx":
        content = _build_xlsx(documents, generated_by, generated_at)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = _build_pdf(documents, generated_by, generated_at)
        content_type = "application/pdf"
    response = HttpResponse(content, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{base_filename}.{file_format}"'
    response["Cache-Control"] = "private, no-store"
    response["X-Content-Type-Options"] = "nosniff"
    return response
